import copy
import csv
import openpyxl
from openpyxl.utils import get_column_letter


def generate_model_template(model):
    return [f for f in model._meta.get_fields() if f.name and not f.null and f.get_internal_type()]


def exportcsv(headers=[], title="Sheet", filename=None, queryset=[], export_csv=False, request=None):
    # Get the totals
    headers_length = len(headers)

    # New workbook
    wb = openpyxl.Workbook()

    # Create a sheet
    sheet = wb.active
    sheet.title = title

    # Set the headers
    for k, col in enumerate(headers):
        cell = sheet.cell(row=1, column=k + 1)
        cell.value = col["name"]

        # Set the size
        if k + 1 <= headers_length:
            sheet.column_dimensions[get_column_letter(k + 1)].width = 20

    # Copy the headers to match the number of fields in data
    fields_length = len([k for k in queryset[0]])
    myheaders = copy.deepcopy(headers[:fields_length])

    # Writing the data
    for i, data in enumerate(queryset):
        # Loop through all the headers
        for j, col in enumerate(myheaders):
            dt = data[col["value"]]
            dat = ",".join(list(dt)) if type(dt) in [list, set] else dt
            sheet.cell(row=i+2, column=j+1).value = dat

    # The output filename
    myfilename = "%s.%s" % (filename, "csv" if export_csv else ".xlsx")

    # Temporary filename for openxl
    temp_filename = "temp_%s" % (myfilename)

    # Temporarily save the file
    if export_csv:
        with open(temp_filename, 'w') as f:
            c = csv.writer(f)
            for r in sheet.rows:
                row_data = []
                for cell in r:
                    row_data.append(cell.value)
                print(row_data)
                c.writerow(row_data)
    else:
        wb.save(temp_filename)

    return temp_filename


if __name__ == '__main__':
    pass
